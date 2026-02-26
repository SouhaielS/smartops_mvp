import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s:%(name)s:%(message)s",
    force=True
)
# src/run_batch.py
from __future__ import annotations

import argparse
import logging
import re
import uuid
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd

# Excel formatting
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Import your extractor (must exist and return a dict with keys:
# PO_Number, Invoice_Number, Invoice_Amount)
from src.extract_invoice import extract_invoice_fields as extract_invoice_dict


# =========================
# CONFIG (defaults)
# =========================
PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
INVOICE_DIR = PROJECT_ROOT / "invoices"
PO_REGISTER_PATH = DATA_DIR / "PO_Register.xlsx"

# Single workbook output (two sheets)
OUT_BATCH_WORKBOOK = DATA_DIR / "Batch_Output.xlsx"

# Required columns in PO register
PO_REQUIRED_COLS = {"PO_Number", "Total_PO_Value", "Amount_Already_Invoiced"}


# =========================
# LOGGING
# =========================
logger = logging.getLogger("run_batch")


# =========================
# HELPERS
# =========================
def _normalize_po(x: Any) -> Optional[str]:
    """
    Normalize PO numbers coming from Excel/extractor:
    - force str
    - strip spaces
    - remove trailing .0 (common when Excel stores as float)
    """
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = re.sub(r"\.0$", "", s)
    return s


def safe_float(x: Any) -> Optional[float]:
    """
    Robust parsing for amounts like:
    '3 748,50'  '3,748.50'  '3.748,50'  '3748.50'  '3748'  'TND 3 748,50'
    Returns None if cannot parse.
    """
    if x is None:
        return None

    s = str(x).strip()
    if s == "":
        return None

    try:
        # keep digits, separators, minus
        s = re.sub(r"[^0-9,.\-]", "", s)

        if s == "" or s in {"-", ".", ","}:
            return None

        # If both separators exist, decide decimal by last occurrence
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                # comma decimal -> remove dots (thousands)
                s = s.replace(".", "").replace(",", ".")
            else:
                # dot decimal -> remove commas (thousands)
                s = s.replace(",", "")
        else:
            # Only one separator type (or none)
            if "," in s and "." not in s:
                # assume comma decimal
                s = s.replace(".", "").replace(",", ".")
            else:
                # dot decimal or none -> remove commas (thousands)
                s = s.replace(",", "")

        return float(s)
    except Exception:
        return None


def _setup_logging(log_level: str = "INFO") -> None:
    logging.basicConfig(
        level=getattr(logging, log_level.upper(), logging.INFO),
        format="%(levelname)s - %(message)s",
    )


# =========================
# DATA MODEL
# =========================
@dataclass
class InvoiceResult:
    batch_id: str
    processed_at: str  # ISO string
    file_name: str
    po_number: Optional[str]
    invoice_number: Optional[str]
    invoice_amount: Optional[float]
    status: str  # VALID | INVALID | OVERBUDGET | PO_NOT_FOUND | NEEDS_REVIEW
    reason: str
    po_budget: Optional[float] = None
    remaining_before: Optional[float] = None
    remaining_after: Optional[float] = None


# =========================
# CONNECT TO EXTRACTOR
# =========================
def extract_invoice_fields(pdf_path: Path) -> Tuple[Optional[str], Optional[str], Optional[float]]:
    """
    Returns: (PO_Number, Invoice_Number, Invoice_Amount)
    """
    data: Dict[str, Any] = extract_invoice_dict(pdf_path) or {}

    po = data.get("PO_Number")
    inv_no = data.get("Invoice_Number")
    amt = data.get("Invoice_Amount")

    po_number = _normalize_po(po)

    invoice_number = None
    if inv_no is not None:
        s = str(inv_no).strip()
        invoice_number = s if s else None

    invoice_amount = safe_float(amt)

    return po_number, invoice_number, invoice_amount


# =========================
# LOAD PO REGISTER (PO LEVEL ONLY)
# =========================
def load_po_register(po_register_path: Path) -> pd.DataFrame:
    df = pd.read_excel(po_register_path)

    missing = [c for c in PO_REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns in PO_Register.xlsx: {missing}. Found: {list(df.columns)}")

    df = df.copy()

    # Normalize PO_Number early (fixes Excel numeric -> "2025003.0" mismatch)
    df["PO_Number"] = df["PO_Number"].apply(_normalize_po)

    # Convert numeric columns robustly
    df["Total_PO_Value"] = df["Total_PO_Value"].apply(safe_float)
    df["Amount_Already_Invoiced"] = df["Amount_Already_Invoiced"].apply(safe_float)

    # Treat missing numeric as 0 for PO math
    df["Total_PO_Value"] = df["Total_PO_Value"].fillna(0.0)
    df["Amount_Already_Invoiced"] = df["Amount_Already_Invoiced"].fillna(0.0)

    # Collapse to ONE row per PO (max budget, sum invoiced)
    grouped = (
        df.groupby("PO_Number", dropna=False)
        .agg({"Total_PO_Value": "max", "Amount_Already_Invoiced": "sum"})
        .reset_index()
    )

    grouped["Remaining_Budget"] = grouped["Total_PO_Value"] - grouped["Amount_Already_Invoiced"]

    # Index by PO_Number for fast lookup
    grouped = grouped.set_index("PO_Number", drop=False)
    return grouped


# =========================
# EXCEL: TABLE + FORMATTING
# =========================
def _apply_excel_table(ws, table_name: str, style_name: str = "TableStyleMedium9") -> None:
    """Turn used range into an Excel Table (pivot-ready)."""
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = ws.dimensions  # e.g. A1:H120
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def _format_sheet_basic(ws, currency_cols=None, header_fill="1F4E79") -> None:
    """Freeze header, header style, widths, currency formats, status conditional formatting."""
    currency_cols = set(currency_cols or [])

    ws.freeze_panes = "A2"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill_obj = PatternFill("solid", fgColor=header_fill)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header styling
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill_obj
        cell.alignment = header_align

    # Auto width + formats
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        header = ws.cell(row=1, column=col).value
        header_str = str(header).strip() if header is not None else ""

        # width
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

        # Currency formatting by header name
        if header_str in currency_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = "#,##0.00"

        # Reason = wrap
        if header_str.lower() == "reason":
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

    # Status conditional formatting
    status_col = None
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col).value).strip().lower() == "status":
            status_col = col
            break

    if status_col:
        col_letter = get_column_letter(status_col)
        rng = f"{col_letter}2:{col_letter}{ws.max_row}"

        fill_green = PatternFill("solid", fgColor="C6EFCE")
        fill_red = PatternFill("solid", fgColor="FFC7CE")
        fill_orange = PatternFill("solid", fgColor="FFEB9C")
        fill_blue = PatternFill("solid", fgColor="D9E1F2")
        fill_purple = PatternFill("solid", fgColor="E6E0FF")  # NEEDS_REVIEW

        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"VALID"'], fill=fill_green)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"INVALID"'], fill=fill_red)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"OVERBUDGET"'], fill=fill_orange)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"PO_NOT_FOUND"'], fill=fill_blue)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"NEEDS_REVIEW"'], fill=fill_purple)
        )


def _add_totals_row_invoice_control(ws) -> None:
    """
    Add Totals row at bottom:
    - Total_All_Amount: sum of invoice_amount (all rows)
    - Total_VALID_Amount: sum of invoice_amount where status == VALID

    Adds two dedicated columns at the end.
    """
    # Map headers -> col index
    col_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(row=1, column=c).value).strip().lower()
        col_map[name] = c

    if "invoice_amount" not in col_map or "status" not in col_map:
        return

    amount_col = col_map["invoice_amount"]
    status_col = col_map["status"]
    last_data_row = ws.max_row
    totals_row = last_data_row + 1

    amount_letter = get_column_letter(amount_col)
    status_letter = get_column_letter(status_col)

    # Add two columns at end (dedicated totals)
    total_all_col = ws.max_column + 1
    total_valid_col = ws.max_column + 2

    ws.cell(row=1, column=total_all_col, value="Total_All_Amount")
    ws.cell(row=1, column=total_valid_col, value="Total_VALID_Amount")

    ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)

    # SUM(all)
    ws.cell(
        row=totals_row,
        column=total_all_col,
        value=f"=SUM({amount_letter}2:{amount_letter}{last_data_row})",
    ).font = Font(bold=True)
    ws.cell(row=totals_row, column=total_all_col).number_format = "#,##0.00"

    # SUMIF(valid)
    ws.cell(
        row=totals_row,
        column=total_valid_col,
        value=(
            f'=SUMIF({status_letter}2:{status_letter}{last_data_row},'
            f'"VALID",{amount_letter}2:{amount_letter}{last_data_row})'
        ),
    ).font = Font(bold=True)
    ws.cell(row=totals_row, column=total_valid_col).number_format = "#,##0.00"

    # Highlight totals row
    fill = PatternFill("solid", fgColor="D9E1F2")
    for col in range(1, ws.max_column + 1):
        ws.cell(row=totals_row, column=col).fill = fill


# =========================
# PROCESS ONE INVOICE
# =========================
def process_invoice(
    df_po: pd.DataFrame,
    pdf_path: Path,
    seen_keys: Set[tuple],
    batch_id: str,
    processed_at_iso: str,
) -> Tuple[InvoiceResult, pd.DataFrame]:
    file_name = pdf_path.name

    # Extract with safety: never crash the whole batch for one PDF
    try:
        po_number, invoice_number, invoice_amount = extract_invoice_fields(pdf_path)
    except Exception as e:
        return (
            InvoiceResult(
                batch_id=batch_id,
                processed_at=processed_at_iso,
                file_name=file_name,
                po_number=None,
                invoice_number=None,
                invoice_amount=None,
                status="INVALID",
                reason=f"Extractor error: {type(e).__name__}: {e}",
            ),
            df_po,
        )

    # ===== RULE A (MVP HARDENING) =====
    # If invoice_number missing => NEEDS_REVIEW, DO NOT consume budget
    if not invoice_number:
        return (
            InvoiceResult(
                batch_id=batch_id,
                processed_at=processed_at_iso,
                file_name=file_name,
                po_number=po_number,
                invoice_number=None,
                invoice_amount=invoice_amount,
                status="NEEDS_REVIEW",
                reason="Invoice number missing",
            ),
            df_po,
        )

    # Duplicate detection (within this run)
    key = (po_number, invoice_number)
    if po_number and invoice_number:
        if key in seen_keys:
            return (
                InvoiceResult(
                    batch_id=batch_id,
                    processed_at=processed_at_iso,
                    file_name=file_name,
                    po_number=po_number,
                    invoice_number=invoice_number,
                    invoice_amount=invoice_amount,
                    status="INVALID",
                    reason="Duplicate invoice detected (same PO_Number + Invoice_Number)",
                ),
                df_po,
            )
        seen_keys.add(key)

    # Basic validations
    if not po_number:
        return (
            InvoiceResult(
                batch_id=batch_id,
                processed_at=processed_at_iso,
                file_name=file_name,
                po_number=None,
                invoice_number=invoice_number,
                invoice_amount=invoice_amount,
                status="INVALID",
                reason="PO missing",
            ),
            df_po,
        )

    if invoice_amount is None or invoice_amount <= 0:
        return (
            InvoiceResult(
                batch_id=batch_id,
                processed_at=processed_at_iso,
                file_name=file_name,
                po_number=po_number,
                invoice_number=invoice_number,
                invoice_amount=invoice_amount,
                status="INVALID",
                reason="Amount missing/zero or not parsed",
            ),
            df_po,
        )

    # PO exists? (fast index lookup)
    if po_number not in df_po.index:
        return (
            InvoiceResult(
                batch_id=batch_id,
                processed_at=processed_at_iso,
                file_name=file_name,
                po_number=po_number,
                invoice_number=invoice_number,
                invoice_amount=invoice_amount,
                status="PO_NOT_FOUND",
                reason="PO not found in register",
            ),
            df_po,
        )

    remaining_before = float(df_po.at[po_number, "Remaining_Budget"])
    po_budget = float(df_po.at[po_number, "Total_PO_Value"])

    # Overbudget?
    if invoice_amount > remaining_before:
        return (
            InvoiceResult(
                batch_id=batch_id,
                processed_at=processed_at_iso,
                file_name=file_name,
                po_number=po_number,
                invoice_number=invoice_number,
                invoice_amount=invoice_amount,
                status="OVERBUDGET",
                reason=f"Invoice exceeds Remaining_Budget ({invoice_amount} > {remaining_before})",
                po_budget=po_budget,
                remaining_before=remaining_before,
                remaining_after=remaining_before,
            ),
            df_po,
        )

    # VALID → update running balance
    new_remaining = remaining_before - invoice_amount
    df_po.at[po_number, "Amount_Already_Invoiced"] = float(df_po.at[po_number, "Amount_Already_Invoiced"]) + invoice_amount
    df_po.at[po_number, "Remaining_Budget"] = new_remaining

    return (
        InvoiceResult(
            batch_id=batch_id,
            processed_at=processed_at_iso,
            file_name=file_name,
            po_number=po_number,
            invoice_number=invoice_number,
            invoice_amount=invoice_amount,
            status="VALID",
            reason="OK",
            po_budget=po_budget,
            remaining_before=remaining_before,
            remaining_after=new_remaining,
        ),
        df_po,
    )


# =========================
# CORE RUNNER (for Streamlit)
# =========================
def run_batch(
    invoice_dir: Path,
    po_register_path: Path,
    out_workbook: Path,
    log_level: str = "INFO",
) -> Path:
    """
    Callable entrypoint for Streamlit (no CLI needed).
    Returns the output workbook path.
    """
    _setup_logging(log_level)

    invoice_dir = Path(invoice_dir)
    po_register_path = Path(po_register_path)
    out_workbook = Path(out_workbook)

    batch_id = uuid.uuid4().hex[:10]
    processed_at_iso = datetime.now().isoformat(timespec="seconds")

    logger.info(f"Batch ID: {batch_id} | Processed at: {processed_at_iso}")
    logger.info(f"Invoice dir: {invoice_dir}")
    logger.info(f"PO register: {po_register_path}")
    logger.info(f"Output workbook: {out_workbook}")

    df_po = load_po_register(po_register_path)

    pdfs = sorted(invoice_dir.glob("*.pdf"))
    if not pdfs:
        raise FileNotFoundError(f"No PDFs found in: {invoice_dir}")

    results: List[InvoiceResult] = []
    seen_keys: Set[tuple] = set()

    for pdf in pdfs:
        logger.info(f"Processing: {pdf.name}")
        res, df_po = process_invoice(df_po, pdf, seen_keys, batch_id, processed_at_iso)
        results.append(res)
        logger.info(f"Status: {res.status} | Reason: {res.reason}")

    # Build output dataframes
    out_df = pd.DataFrame([asdict(r) for r in results])

    # Write ONE workbook with TWO sheets + totals row + formatting + Excel tables
    out_workbook.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_workbook, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="Invoice_Control", index=False)

        # Keep PO sheet always updated with ONLY VALID invoices consumed
        df_po_out = df_po.reset_index(drop=True)
        df_po_out.to_excel(writer, sheet_name="PO_Register_Updated", index=False)

        ws_inv = writer.book["Invoice_Control"]
        ws_po = writer.book["PO_Register_Updated"]

        # Totals row (adds Total_All_Amount + Total_VALID_Amount columns at end)
        _add_totals_row_invoice_control(ws_inv)

        # Formatting
        _format_sheet_basic(
            ws_inv,
            currency_cols={
                "invoice_amount",
                "po_budget",
                "remaining_before",
                "remaining_after",
                "Total_All_Amount",
                "Total_VALID_Amount",
            },
            header_fill="1F4E79",
        )
        _format_sheet_basic(
            ws_po,
            currency_cols={"Total_PO_Value", "Amount_Already_Invoiced", "Remaining_Budget"},
            header_fill="2F5597",
        )

        # Pivot-ready tables
        _apply_excel_table(ws_inv, table_name="InvoiceControlTable", style_name="TableStyleMedium9")
        _apply_excel_table(ws_po, table_name="PORegisterTable", style_name="TableStyleMedium9")

    return out_workbook


# =========================
# CLI
# =========================
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Invoice Control + PO Budget Consumption Batch")
    p.add_argument("--invoice-dir", type=str, default=str(INVOICE_DIR), help="Directory containing invoice PDFs")
    p.add_argument("--po-register", type=str, default=str(PO_REGISTER_PATH), help="Path to PO_Register.xlsx")
    p.add_argument("--out", type=str, default=str(OUT_BATCH_WORKBOOK), help="Output Excel workbook path")
    p.add_argument("--log-level", type=str, default="INFO", help="DEBUG/INFO/WARNING/ERROR")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    out_path = run_batch(
        invoice_dir=Path(args.invoice_dir),
        po_register_path=Path(args.po_register),
        out_workbook=Path(args.out),
        log_level=args.log_level,
    )

    logger.info(f"✅ Batch workbook written to: {out_path}")


if __name__ == "__main__":
    main()